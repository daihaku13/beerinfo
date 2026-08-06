# -*- coding: utf-8 -*-
"""
step3_excel_writer.py
================
⑤ Python（Excel書き込み）
既存のExcelテンプレート(BeerInfo-template1.xlsm)をもとに、管理番号ごとに
シートを複製し、決まったセル位置へ統合済み結果を書き込む。

全件処理が終わった後に1回だけ呼び出す想定:

    from step3_excel_writer import write_to_excel
    output_path = write_to_excel(consolidated_results)

■セルマッピング(BeerInfo-template.xlsm「template」シート実物より確定／2026年08月改訂)
    B2  ：管理番号
    B3  ：ビア名称
    B5  ：紹介１（結合 B5:B8）
    B9  ：紹介２（結合 B9:B12）
    B13 ：紹介３（結合 B13:B16）
    B18 ：開催期間
    B19 ：営業時間
    B20 ：定休日
    B21 ：雨天営業
    B22 ：ビアタイプ
    B23 ：概要
    B24 ：システム（結合 B24:B25）
    B26 ：料金(税込)（結合 B26:B30）
    E2  ：公式URL
    E3  ：ビア紹介URL
    E5  ：料理情報（結合 E5:E16）
    E17 ：ドリンク（結合 E17:E26）
    E28 ：ポイント１
    E29 ：ポイント２
    E30 ：ポイント３

    ※本テンプレートにはVBAマクロ(AI_CHECK指令出力／BeerDB_Update)が組み込まれている。
      本モジュールはkeep_vba=Trueで読み書きするため、これらのマクロ自体は保持されるが、
      本モジュール(Pythonパイプライン)からは呼び出さない。利用者がExcelを開いて手動で
      実行する用途のマクロであり、自動化フローとは独立した機能として扱う。
    ※旧マッピング(B3:管理番号/B4:ビア名称/B6:紹介１/B9:紹介２/B12:紹介３/B17:開催期間
      /B18:営業時間/B19:定休日/B20:雨天営業/B21:ビアタイプ/B22:概要/B23:システム/
      B24:料金/E3:公式URL/E4:ビア紹介URL/E7:料理/E11:ドリンク/E17-19:ポイント1-3)は
      旧テンプレート用。新テンプレートでは全面的に位置が変更されているため注意。
"""

import re
import shutil
import uuid
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# 設定
# ---------------------------------------------------------------------------

# テンプレートファイルのパス・シート名は環境に合わせて変更してください。
TEMPLATE_PATH = "/var/www/beerinfo/templates/BeerInfo-template.xlsm"
TEMPLATE_SHEET_NAME = "template"

EXCEL_OUTPUT_DIR = Path("/var/www/beerinfo/excel_output")
EXCEL_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Excel作成後、OneDrive同期フォルダへも自動コピーする
ONEDRIVE_DIR = Path("/root/onedrive")

# 統合済みdictのキー → 書き込み先セル(結合範囲の場合は左上セル)
# 2026年08月改訂：新テンプレート(BeerInfo-template.xlsm)のセル配置に合わせて全面更新
CELL_MAP = {
    "管理番号": "B2",
    "ビア名称": "B3",
    "紹介１": "B5",
    "紹介２": "B9",
    "紹介３": "B13",
    "開催期間": "B18",
    "営業時間": "B19",
    "定休日": "B20",
    "雨天営業": "B21",
    "ビアタイプ": "B22",
    "概要": "B23",
    "システム": "B24",
    "料金(税込)": "B26",
    "公式URL": "E2",
    "ビア紹介URL": "E3",
    "料理情報": "E5",
    "ドリンク": "E17",
    "ポイント１": "E28",
    "ポイント２": "E29",
    "ポイント３": "E30",
}

INVALID_SHEET_NAME_CHARS = r'[\\/*?:\[\]]'


def _safe_sheet_name(id, name):
    """Excelのシート名として使える文字列に整形する(31文字制限・使用不可文字の除去)"""
    raw = f"{id}_{name}" if name else str(id)
    cleaned = re.sub(INVALID_SHEET_NAME_CHARS, "_", raw)
    return cleaned[:31]


def write_to_excel(consolidated_results: list, template_path: str = TEMPLATE_PATH) -> dict:
    """
    統合済み結果のリストを受け取り、テンプレートを管理番号ごとに複製して
    セルマッピングに従って値を書き込み、1つのExcelファイルとして保存する。
    保存後、OneDrive同期フォルダ(/root/onedrive)へもコピーする。

    Parameters
    ----------
    consolidated_results : list[dict]
        storage.build_consolidated_dict() / save_consolidated_result() が返す
        dictのリスト(全件処理が終わった後にまとめて渡す)
    template_path : str
        テンプレートファイル(.xlsm)のパス

    Returns
    -------
    dict
        {
          "excel_path": "保存したExcelファイルのパス",
          "onedrive_path": "OneDriveへコピーしたパス(コピー失敗時はNone)",
          "onedrive_error": "コピーに失敗した場合のエラーメッセージ(成功時はNone)"
        }
    """
    if not consolidated_results:
        raise ValueError("consolidated_results が空です。書き込む対象がありません。")

    # マクロ付き(.xlsm)を壊さないよう keep_vba=True で読み込む
    wb = load_workbook(template_path, keep_vba=True)

    if TEMPLATE_SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"テンプレート内にシート「{TEMPLATE_SHEET_NAME}」が見つかりません。"
            f"存在するシート: {wb.sheetnames}"
        )

    template_ws = wb[TEMPLATE_SHEET_NAME]

    used_sheet_names = set(wb.sheetnames)

    for consolidated in consolidated_results:
        id = consolidated.get("管理番号", "")
        name = consolidated.get("ビア名称", "")

        # テンプレートシートを複製
        new_ws = wb.copy_worksheet(template_ws)

        # シート名を決定(重複回避)
        base_name = _safe_sheet_name(id, name)
        sheet_name = base_name
        suffix = 1
        while sheet_name in used_sheet_names:
            suffix += 1
            sheet_name = f"{base_name[:28]}_{suffix}"
        used_sheet_names.add(sheet_name)
        new_ws.title = sheet_name

        # セルマッピングに従って値を書き込む(結合セルは左上セルのみ書き込めばよい)
        for key, cell_address in CELL_MAP.items():
            value = consolidated.get(key, "")
            new_ws[cell_address] = value

    # 元のtemplateシートはひな形として残しておく(削除したい場合は下記を有効化)
    # wb.remove(template_ws)

    # 秒精度のタイムスタンプだけだと、複数実行が同じ秒に完了した場合に
    # ファイル名が衝突し、片方の結果が上書きされてしまうため、
    # 短い一意な識別子(uuid4の先頭8文字)を付加して衝突を防ぐ
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    unique_suffix = uuid.uuid4().hex[:8]
    output_path = EXCEL_OUTPUT_DIR / f"BeerInfo_{timestamp}_{unique_suffix}.xlsm"
    wb.save(output_path)

    # OneDrive同期フォルダへコピー(失敗してもExcel本体の作成自体は成功扱いとする)
    onedrive_path = None
    onedrive_error = None
    try:
        ONEDRIVE_DIR.mkdir(parents=True, exist_ok=True)
        dest = ONEDRIVE_DIR / output_path.name
        shutil.copy2(output_path, dest)
        onedrive_path = str(dest)
    except Exception as e:
        onedrive_error = str(e)

    return {
        "excel_path": str(output_path),
        "onedrive_path": onedrive_path,
        "onedrive_error": onedrive_error,
    }
